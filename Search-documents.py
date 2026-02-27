# ============================================================
# 1. IMPORTS
# ============================================================

import os
import sys
import csv
import hashlib
import unicodedata
import shutil
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

# ============================================================
# 3. UTILITY FUNCTIONS
# ============================================================

def clean_filename(name: str) -> str:
    """
    Normalize filename to a safe, comparable form.
    """
    if not isinstance(name, str):
        name = str(name)
    name = unicodedata.normalize("NFC", name)
    return name.strip()


def normalize_path(path: str) -> str:
    """
    Normalize paths for consistent comparison.
    """
    return os.path.normpath(os.path.abspath(path))


def file_hash(path: str) -> str:
    """
    Compute SHA-256 hash of a file.
    """
    h = hashlib.sha256()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
    except Exception as e:
        diag(f"Hash failed for: {path} ({e})")
        return "None""

    return h.hexdigest()


def write_log(msg: str) -> None:
    """
    Append a line to the deletion/quarantine log.
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}\n"
    try:
        with open(LOG_FILE, "a", encoding="utf-8", errors="replace") as f:
            f.write(line)
    except Exception:
        # Logging must never crash the app
        pass


# ============================================================
# 4. FOLDER SCANNING
# ============================================================

def scan_folder(folder, multi=False):
    results = {}

    for root, dirs, files in os.walk(folder):
        for filename in files:
            full_path = os.path.join(root, filename)

            try:
                size = os.path.getsize(full_path)
                mtime = os.path.getmtime(full_path)
            except Exception as e:
                diag(f"SCAN: Skipping unreadable file: {full_path} ({e})")
                continue

            if multi:
                # MULTI-MATCH MODE (folderB)
                results.setdefault(filename, []).append((full_path, size, mtime))
                diag(f"SCAN: Adding B candidate for {filename}: {full_path}")
            else:
                # SINGLE-MATCH MODE (folderA)
                results[filename] = (full_path, size, mtime)
                diag(f"SCAN: Adding A file {filename}: {full_path}")
   
    return results

def dump_scan_results(filesA, filesB):
    print("\n================ SCAN RESULTS ================")

    print("\n--- FILES A ---")
    for key, value in filesA.items():
        print(f"  A key: {key}")
        print(f"    {value}")

    print("\n--- FILES B ---")
    for key, value in filesB.items():
        print(f"  B key: {key}")
        for entry in value:
            print(f"    {entry}")

    print("\n==============================================\n")
# ============================================================
# 5. COMPARISON ENGINE
# ============================================================

# ============================================================
# Test Harness for matches, mismatched and missing
# ============================================================
def dump_diagnostics(matches, mismatched, missing):
    print("\n================ DIAGNOSTIC DUMP ================")

    print("\n--- MATCHES ---")
    if not matches:
        print("  (none)")
    else:
        for rel, pathA, pathB in matches:
            print(f"  MATCH: {rel}")
            print(f"    A: {pathA}")
            print(f"    B: {pathB}")

    print("\n--- MISMATCHED ---")
    if not mismatched:
        print("  (none)")
    else:
        for rel, pathA, b_list in mismatched:
            print(f"  MISMATCH: {rel}")
            print(f"    A: {pathA}")
            for pathB, sizeB, mtimeB in b_list:
                print(f"    B: {pathB} (size={sizeB}, mtime={mtimeB})")

    print("\n--- MISSING ---")
    if not missing:
        print("  (none)")
    else:
        for rel, pathA in missing:
            print(f"  MISSING: {rel}")
            print(f"    A: {pathA}")

    print("\n=================================================\n")


def compare_folders_recursive(folderA, folderB,
                              output_dir,
                              progress_callback=None,
                              status_callback=None):

    import os
    from datetime import datetime

    import config

    from utils.output import create_timestamped_folder
    
    # Apply timestamped output behavior ONCE
    if config.TIMESTAMPED_OUTPUT:
        output_dir = create_timestamped_folder(output_dir)

    if status_callback:
        status_callback("Scanning folders...")

    # Results from SCAN
    filesA = scan_folder(folderA, multi=False)
    filesB = scan_folder(folderB, multi=True)
    
    diag(f"SCAN: FolderA files = {len(filesA)}")
    diag(f"SCAN: FolderB files = {len(filesB)}")


    dump_scan_results(filesA, filesB)


    if progress_callback:
        progress_callback(20)

    matches = []
    mismatched = []
    missing = []

    # ------------------------------------------------------------
    # Main comparison loop
    # ------------------------------------------------------------
    for rel, (pathA, sizeA, mtimeA) in filesA.items():

        candidatesB = filesB.get(rel, [])
        if not candidatesB:
            missing.append((rel, pathA))
            continue

        found_match = False
        mismatch_list = []

        for pathB, sizeB, mtimeB in candidatesB:
            diag(f"COMPARE: {rel}")

            # TIMESTAMP + SIZE MATCH
            if sizeA == sizeB and abs(mtimeA - mtimeB) <= 1.0:
                diag(f"MATCH: {rel} (timestamp+size)")

                matches.append((rel, pathA, pathB))
                found_match = True
                # DO NOT break — allow multi-match
                continue

            # ACCURATE MODE (hash confirm)
            if sizeA == sizeB:
                diag(f"MATCH: {rel} (hash)")

                hashA = file_hash(pathA)
                hashB = file_hash(pathB)
                if hashA == hashB:
                    matches.append((rel, pathA, pathB))
                    found_match = True
                    continue

            # HASH-ONLY MODE
            if config.HASH_ONLY_MODE:
                hashA = file_hash(pathA)
                hashB = file_hash(pathB)
                if hashA == hashB:
                    matches.append((rel, pathA, pathB))
                    found_match = True
                    continue

            # Otherwise mismatch
            diag(f"MISMATCH: {rel} vs {pathB}")
            mismatch_list.append((pathB, sizeB, mtimeB))

        # FINAL DECISION
        if found_match:
            # We DO NOT suppress mismatches anymore.
            # Mixed cases should show both matches and mismatches.
            if mismatch_list:
                mismatched.append((rel, pathA, mismatch_list))
        else:
            # No matches at all
            diag(f"MISSING: {rel}")

            if mismatch_list:
                mismatched.append((rel, pathA, mismatch_list))
            else:
                missing.append((rel, pathA))

    if progress_callback:
        progress_callback(50)

    # ------------------------------------------------------------
    # DIAGNOSTIC DUMP
    # ------------------------------------------------------------
    dump_diagnostics(matches, mismatched, missing)

    # ------------------------------------------------------------
    # Write reports
    # ------------------------------------------------------------
    if status_callback:
        status_callback("Writing reports...")

    diag("WRITING: Starting report generation")

    write_all_reports(
        output_dir,
        matches,
        mismatched,
        missing,
        status_callback=status_callback,
        progress_callback=progress_callback
    )
    diag("WRITING: Reports complete")

    if status_callback:
        status_callback("Comparison complete.")

    return matches, mismatched, missing

# ============================================================
# 6. OUTPUT WRITERS
# ============================================================
def print_summary(matches, mismatched, missing, status_callback=None):
    total_matches = len(matches)
    total_mismatches = len(mismatched)
    total_missing = len(missing)

    from collections import Counter
    match_names = [name for name, _, _ in matches]
    multi_match_counts = Counter(match_names)
    multi_match_cases = sum(1 for k, v in multi_match_counts.items() if v > 1)

    mismatch_names = {name for name, _, _ in mismatched}
    mixed_cases = sum(1 for name in mismatch_names if name in multi_match_counts)

    summary = [
        "=============== SUMMARY REPORT ===============",
        f"Total Exact Matches:     {total_matches}",
        f"Total Mismatches:        {total_mismatches}",
        f"Total Missing Files:     {total_missing}",
        "",
        f"Multi-Match Cases:       {multi_match_cases}",
        f"Mixed Match/Mismatch:    {mixed_cases}",
        "=============================================="
    ]

    summary_text = "\n".join(summary)

    print(summary_text)
    if status_callback:
        status_callback(summary_text)

    return summary_text
    
def write_summary_output(output_dir, matches, mismatched, missing,
                         progress_callback=None, status_callback=None):

    
    import os
        
    summary_path = os.path.join(output_dir, "summary.txt")

    if status_callback:
        status_callback(f"Writing Summary report to: {summary_path}")
    if progress_callback:
        progress_callback(95)

    # Build summary text using the same logic as print_summary
    total_matches = len(matches)
    total_mismatches = len(mismatched)
    total_missing = len(missing)

    from collections import Counter
    match_names = [name for name, _, _ in matches]
    multi_match_counts = Counter(match_names)
    multi_match_cases = sum(1 for k, v in multi_match_counts.items() if v > 1)

    mismatch_names = {name for name, _, _ in mismatched}
    mixed_cases = sum(1 for name in mismatch_names if name in multi_match_counts)

    lines = []
    lines.append("=============== SUMMARY REPORT ===============")
    lines.append(f"Total Exact Matches:     {total_matches}")
    lines.append(f"Total Mismatches:        {total_mismatches}")
    lines.append(f"Total Missing Files:     {total_missing}")
    lines.append("")
    lines.append(f"Multi-Match Cases:       {multi_match_cases}")
    lines.append(f"Mixed Match/Mismatch:    {mixed_cases}")
    lines.append("==============================================")

    summary_text = "\n".join(lines)

    with open(summary_path, "w", encoding="utf-8", errors="replace") as f:
        f.write(summary_text)

    if status_callback:
        status_callback("Summary report written.")

def write_excel_output(output_dir, matches, mismatched, missing,
                       progress_callback=None, status_callback=None):

    import os
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill


    excel_path = os.path.join(output_dir, "comparison.xlsx")

    # Optional: writer-level status update
    if status_callback:
        status_callback(f"Writing Excel report to: {excel_path}")
    if progress_callback:
        progress_callback(60)
    diag("ENTERING EXCEL WRITER")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"

        header = [
            "Status",
            "Filename",
            "Path A",
            "Timestamp A",
            "Path B",
            "Timestamp B",
        ]
        ws.append(header)

        diag("Checkpoint Exact Matches")

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # ------------------------------------------------------------
        # Exact Matches
        # ------------------------------------------------------------
        for name, pathA, pathB in matches:

            if not pathA or not os.path.exists(pathA):
                diag(f"Excel: Skipping invalid pathA for match: {pathA}")
                continue
            if not pathB or not os.path.exists(pathB):
                diag(f"Excel: Skipping invalid pathB for match: {pathB}")
                continue

            try:
                tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsA = ""
            try:
                tsB = datetime.fromtimestamp(os.path.getmtime(pathB)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsB = ""

            row = ["Exact Match", name, pathA, tsA, pathB, tsB]
            ws.append(row)

            for cell in ws[ws.max_row]:
                cell.fill = green

        # ------------------------------------------------------------
        # Mismatches
        # ------------------------------------------------------------
        diag("Checkpoint Before Mismatches")

        for name, pathA, b_list in mismatched:

            if not pathA or not os.path.exists(pathA):
                diag(f"Excel: Skipping invalid pathA for mismatch: {pathA}")
                continue

            try:
                tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsA = ""

            for pathB, _, _ in b_list:

                if not pathB or not os.path.exists(pathB):
                    diag(f"Excel: Skipping invalid pathB for mismatch: {pathB}")
                    continue

                try:
                    tsB = datetime.fromtimestamp(os.path.getmtime(pathB)).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tsB = ""

                row = ["Mismatch", name, pathA, tsA, pathB, tsB]
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.fill = yellow

        # ------------------------------------------------------------
        # Missing Files
        # ------------------------------------------------------------
        diag("Checkpoint Before Missing Files")

        for name, pathA in missing:

            if not pathA or not os.path.exists(pathA):
                diag(f"Excel: Skipping invalid pathA for missing: {pathA}")
                continue

            try:
                tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsA = ""

            row = ["Missing in Folder B", name, pathA, tsA, "", ""]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = red
        diag("Checkpoint Before Save")

        wb.save(excel_path)
        diag("Checkpoint After Save")
    except Exception as e:
        diag(f"ERROR writing Excel report: {e}")
        print(f"ERROR writing Excel report: {e}")



def write_csv_output(output_dir, matches, mismatched, missing,
                       progress_callback=None, status_callback=None):

    import os
    import csv
    from datetime import datetime
    
    csv_path = os.path.join(output_dir, "comparison.csv")
    
    if status_callback:
        status_callback(f"Writing CSV report to: {csv_path}")
    if progress_callback:
        progress_callback(75)

    try:
        with open(csv_path, "w", newline="", encoding="utf-8", errors="replace") as f:
            writer = csv.writer(f)
            writer.writerow(["Status", "Filename", "Path A", "Timestamp A", "Path B", "Timestamp B"])

            # ------------------------------------------------------------
            # Exact Matches
            # ------------------------------------------------------------
            for name, pathA, pathB in matches:

                if not pathA or not os.path.exists(pathA):
                    diag(f"CSV: Skipping invalid pathA for match: {pathA}")
                    continue
                if not pathB or not os.path.exists(pathB):
                    diag(f"CSV: Skipping invalid pathB for match: {pathB}")
                    continue

                try:
                    tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tsA = ""
                try:
                    tsB = datetime.fromtimestamp(os.path.getmtime(pathB)).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tsB = ""

                writer.writerow(["Exact Match", name, pathA, tsA, pathB, tsB])

            # ------------------------------------------------------------
            # Mismatches
            # ------------------------------------------------------------
            for name, pathA, b_list in mismatched:

                if not pathA or not os.path.exists(pathA):
                    diag(f"CSV: Skipping invalid pathA for mismatch: {pathA}")
                    continue

                try:
                    tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tsA = ""

                for pathB, _, _ in b_list:

                    if not pathB or not os.path.exists(pathB):
                        diag(f"CSV: Skipping invalid pathB for mismatch: {pathB}")
                        continue

                    try:
                        tsB = datetime.fromtimestamp(os.path.getmtime(pathB)).strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        tsB = ""

                    writer.writerow(["Mismatch", name, pathA, tsA, pathB, tsB])

            # ------------------------------------------------------------
            # Missing Files
            # ------------------------------------------------------------
            for name, pathA in missing:

                if not pathA or not os.path.exists(pathA):
                    diag(f"CSV: Skipping invalid pathA for missing: {pathA}")
                    continue

                try:
                    tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tsA = ""

                writer.writerow(["Missing in Folder B", name, pathA, tsA, "", ""])
    except Exception as e:
        diag(f"ERROR writing CSV report: {e}")
        print(f"ERROR writing CSV report: {e}")


def write_text_output(output_dir, matches, mismatched, missing,
                       progress_callback=None, status_callback=None):

    import os
    from datetime import datetime
    
    txt_path = os.path.join(output_dir, "comparison.txt")
    
    # Optional: writer-level status update
    if status_callback:
        status_callback(f"Writing Text report to: {txt_path}")
    if progress_callback:
        progress_callback(85)

    try:
        with open(txt_path, "w", encoding="utf-8", errors="replace") as f:

            # ------------------------------------------------------------
            # Exact Matches
            # ------------------------------------------------------------
            f.write("=== Exact Matches ===\n")

            for name, pathA, pathB in matches:

                if not pathA or not os.path.exists(pathA):
                    diag(f"Text: Skipping invalid pathA for match: {pathA}")
                    continue
                if not pathB or not os.path.exists(pathB):
                    diag(f"Text: Skipping invalid pathB for match: {pathB}")
                    continue

                f.write(f"{name}\n")
                f.write(f"  A: {pathA}\n")
                f.write(f"  B: {pathB}\n\n")

            # ------------------------------------------------------------
            # Mismatches
            # ------------------------------------------------------------
            f.write("=== Mismatches ===\n")

            for name, pathA, b_list in mismatched:

                if not pathA or not os.path.exists(pathA):
                    diag(f"Text: Skipping invalid pathA for mismatch: {pathA}")
                    continue

                try:
                    tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tsA = ""

                f.write(f"{name}\n")
                f.write(f"  A: {pathA} (ts={tsA})\n")

                for pathB, sizeB, mtimeB in b_list:

                    if not pathB or not os.path.exists(pathB):
                        diag(f"Text: Skipping invalid pathB for mismatch: {pathB}")
                        continue

                    try:
                        tsB = datetime.fromtimestamp(mtimeB).strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        tsB = ""

                    f.write(f"  B: {pathB} (size={sizeB}, ts={tsB})\n")

                f.write("\n")

            # ------------------------------------------------------------
            # Missing Files
            # ------------------------------------------------------------
            f.write("=== Missing in Folder B ===\n")

            for name, pathA in missing:

                if not pathA or not os.path.exists(pathA):
                    diag(f"Text: Skipping invalid pathA for missing: {pathA}")
                    continue

                f.write(f"{name}\n")
                f.write(f"  A: {pathA}\n\n")
    except Exception as e:
        diag(f"ERROR writing Text report: {e}")
        print(f"ERROR writing Text report: {e}")


def write_all_reports(output_dir, matches, mismatched, missing,
                      status_callback=None, progress_callback=None):
    """
    Unified wrapper that writes Excel, CSV, and Text reports.
    Ensures each writer is isolated so one failure does not break the others.
    Provides consistent diagnostic output and callback handling.
    """
    import os
    from utils.output import create_timestamped_folder  # if needed
 
    # Ensure output directory exists
    try:
        os.makedirs(output_dir, exist_ok=True)
    except Exception as e:
        diag(f"ERROR: Could not create output directory: {output_dir} ({e})")
        if status_callback:
            status_callback(f"ERROR: Could not create output directory: {output_dir}")
        return

    # ------------------------------------------------------------
    # Excel Writer
    # ------------------------------------------------------------
    try:
        write_excel_output(
            output_dir,
            matches,
            mismatched,
            missing,
            progress_callback=progress_callback,
            status_callback=status_callback
        )
        diag(f"Excel writer completed")
    except Exception as e:
        diag(f"ERROR writing Excel report: {e}")
        if status_callback:
            status_callback(f"ERROR writing Excel report: {e}")

    # ------------------------------------------------------------
    # CSV Writer
    # ------------------------------------------------------------
    try:
        write_csv_output(
            output_dir,
            matches,
            mismatched,
            missing,
            progress_callback=progress_callback,
            status_callback=status_callback
        )
        diag(f"CSV writer completed")
    except Exception as e:
        diag(f"ERROR writing CSV report: {e}")
        if status_callback:
            status_callback(f"ERROR writing CSV report: {e}")

    # ------------------------------------------------------------
    # Text Writer
    # ------------------------------------------------------------
    try:
        write_text_output(
            output_dir,
            matches,
            mismatched,
            missing,
            progress_callback=progress_callback,
            status_callback=status_callback
        )
        diag(f"Text writer completed")
    except Exception as e:
        diag(f"ERROR writing Text report: {e}")
        if status_callback:
            status_callback(f"ERROR writing Text report: {e}")
            
    
    # Write summary.txt
    write_summary_output(output_dir, matches, mismatched, missing,
                     progress_callback, status_callback)

    # Final status
    if status_callback:
        status_callback(f"Reports written to: {output_dir}")
# ============================================================
# 7. DELETION / QUARANTINE ENGINE
# ============================================================

def ensure_quarantine() -> str:
    """
    Ensure a quarantine folder exists next to the EXE/script.
    """
    base = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else __file__)
    q = os.path.join(base, "Quarantine")
    os.makedirs(q, exist_ok=True)
    return q


def _delete_or_quarantine(path: str, reason: str):
    """
    Delete or move a file based on global flags.
    """
    if DRY_RUN:
        write_log(f"[DRY-RUN] Would handle: {path} ({reason})")
        return

    if USE_QUARANTINE:
        q = ensure_quarantine()
        base = os.path.basename(path)
        target = os.path.join(q, base)
        try:
            shutil.move(path, target)
            action = f"QUARANTINED -> {target}"
        except Exception as e:
            write_log(f"[ERROR] Failed to quarantine {path}: {e}")
            return
    else:
        try:
            os.remove(path)
            action = "DELETED"
        except Exception as e:
            write_log(f"[ERROR] Failed to delete {path}: {e}")
            return

    write_log(f"[{action}] {path} ({reason})")


def process_deletions(matches, mismatched):
    """
    Apply deletion/quarantine rules based on global flags.
    """
    
    # Exact matches
    if DELETE_EXACT_MATCHES:
        for _, _, pathB in matches:
            _delete_or_quarantine(pathB, "Exact Match")

    # Mismatch candidates
    if DELETE_CANDIDATES:
        for _, _, b_list in mismatched:
            for pathB, _, _ in b_list:
                _delete_or_quarantine(pathB, "Mismatch Candidate")


# ============================================================
# 8. GUI LAUNCHER
# ============================================================

def launch_gui():
        
    # Handle splash (PyInstaller)
    pyi_splash = None
    if getattr(sys, 'frozen', False):
        try:
            import pyi_splash
            pyi_splash.update_text("Loading GUI...")
        except Exception:
            pyi_splash = None

    root = tk.Tk()
    root.title("Reconciliation Tool")
    root.geometry("800x800")
    root.resizable(False, False)

    # -----------------------------
    # Folder Selection
    # -----------------------------
    folderA_var = tk.StringVar()
    folderB_var = tk.StringVar()
    output_dir_var = tk.StringVar()

    compare_mode_var = tk.StringVar(value="timestamp")  # default mode

    findall_var = tk.BooleanVar()
    dryrun_var = tk.BooleanVar(value=True)
    deletematches_var = tk.BooleanVar()
    deletecandidates_var = tk.BooleanVar()
    quarantine_var = tk.BooleanVar(value=True)
    diagnostic_var = tk.BooleanVar()# Future if arg should be available on UI
    
    def browse_folderA():
        path = filedialog.askdirectory()
        if path:
            folderA_var.set(path)

    def browse_folderB():
        path = filedialog.askdirectory()
        if path:
            folderB_var.set(path)
         
    # Output Path         
    tk.Label(root, text="Output Folder:", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20)
    tk.Entry(root, textvariable=output_dir_var, width=60).pack(anchor="w", padx=20)
    
    diag(f"GUI assigning variable to UI")

    def browse_output_dir():
        path = filedialog.askdirectory()
        if path:
            output_dir_var.set(path)

    tk.Button(root, text="Browse", command=browse_output_dir).pack(anchor="w", padx=20, pady=(0, 20))
 
    tk.Label(root, text="Folder A (Source):", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20, pady=(20, 0))
    tk.Entry(root, textvariable=folderA_var, width=60).pack(anchor="w", padx=20)
    tk.Button(root, text="Browse", command=browse_folderA).pack(anchor="w", padx=20, pady=(0, 10))
   
    tk.Label(root, text="Folder B (Target):", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20)
    tk.Entry(root, textvariable=folderB_var, width=60).pack(anchor="w", padx=20)
    tk.Button(root, text="Browse", command=browse_folderB).pack(anchor="w", padx=20, pady=(0, 20))
   
    # -----------------------------
    # Comparison Options
    # -----------------------------
    tk.Label(root, text="Comparison Options:", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20)
   
    tk.Radiobutton(root, text="Timestamp + Size (Less Accurate Fastest)", variable=compare_mode_var, value="timestamp").pack(anchor="w", padx=40)
    tk.Radiobutton(root, text="Hash Comparison (Accurate Slower)", variable=compare_mode_var, value="hash").pack(anchor="w", padx=40)
    tk.Radiobutton(root, text="Hash-Only Mode (Accurate Faster)", variable=compare_mode_var, value="hashonly").pack(anchor="w", padx=40)
   
    find_all_var = tk.BooleanVar()
    tk.Label(root, text="Output Options:", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20)
    tk.Checkbutton(root, text="Find All Locations in Folder B", variable= find_all_var).pack(anchor="w", padx=40, pady=(0, 10))
   
    # -----------------------------
    # Deletion / Quarantine Options
    # -----------------------------
    tk.Label(root, text="Cleanup Options:", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20)

    tk.Checkbutton(root, text="Dry Run (No Changes Made)", variable=dryrun_var).pack(anchor="w", padx=40)
    
    tk.Checkbutton(root, text="Delete Exact Matches", variable=deletematches_var).pack(anchor="w", padx=40)
    
    tk.Checkbutton(root, text="Delete Mismatch Candidates", variable=deletecandidates_var).pack(anchor="w", padx=40)
    
    tk.Checkbutton(root, text="Use Quarantine Folder", variable=quarantine_var).pack(anchor="w", padx=40)
    
    
    # -----------------------------
    # Progress Bar + Status Label
    # -----------------------------
    status_var = tk.StringVar(value="Ready")
    tk.Label(root,textvariable=status_var,font=("Segoe UI", 9),anchor="w").pack(fill="x", padx=20, pady=(10, 0))

    progress = ttk.Progressbar(root, orient="horizontal", length=500, mode="determinate")
    progress.pack(padx=20, pady=(0, 10))

    # -----------------------------
    # Progress Bar Helper Functions
    # -----------------------------
    def set_status(msg):
        status_var.set(msg)
        root.update_idletasks()

    def set_progress(value, maximum=100):
        progress["maximum"] = maximum
        progress["value"] = value
        root.update_idletasks()
    # -----------------------------
    # Set arguments 
    # -----------------------------
    
    import config
    config.initialize_runtime()
    
    diag(f"GUI entering run")

    # -----------------------------
    # Run Button
    # -----------------------------
    def run_comparison():
        diag(f"Run button clicked")

        # get data from UI varables
        folderA = folderA_var.get()
        folderB = folderB_var.get()
        output_dir = output_dir_var.get()
        mode = compare_mode_var.get()

        # Reset flags
        config.initialize_runtime()
        config.HASH_ONLY_MODE = False
        config.HASH_COMPARE_MODE = False
        config.FIND_ALL_LOCATIONS_MODE = False

        # Apply mode
        config.HASH_ONLY_MODE = (mode == "hashonly")
        config.HASH_COMPARE_MODE = (mode == "hash")
        config.HASH_COMPARE_MODE = (mode == "timestamp") #default 
        
        # Apply checkboxes
        config.FIND_ALL_LOCATIONS_MODE = findall_var.get()
        diag(f"Before Setting DRY_RUN {dryrun.var.get()}")
        config.DRY_RUN = dryrun_var.get()
        diag(f"After Setting DRY_RUN {dryrun.var.get()}")
        config.DELETE_EXACT_MATCHES = deletematches_var.get()
        config.DELETE_CANDIDATES = deletecandidates_var.get()
        config.USE_QUARANTINE = quarantine_var.get()

        if not folderA or not folderB:
            messagebox.showerror("Error", "Please select both Folder A and Folder B.")
            return

        if not os.path.isdir(folderA) or not os.path.isdir(folderB):
            messagebox.showerror("Error", "One or both selected folders are invalid.")
            return

        diag(f"Variables from GUI")
        diag(f"GUI initialized - Folder A {folderA}")
        diag(f"GUI initialized - Folder B {folderB}")
        diag(f"GUI initialized - Output  {output_dir}")
        diag(f"GUI initialized - Mode  {mode} ")
        diag(f"GUI initialized - DryRun {DRY_RUN} ")
        diag(f"GUI initialized - DELETE EXACT MATCHES {DELETE_EXACT_MATCHES}")
        diag(f"GUI initialized - DELETE CANDIDATES {DELETE_CANDIDATES}")
        diag(f"GUI initialized - DELETE Use Quarantine {USE_QUARANTINE}")
        #diag(f"GUI initialized - DELETE Diagnostic mode {DIAGNOSTIC_MODE}")


    # -----------------------------  
    # Initialize the progress bar
    # -----------------------------
        set_status("Starting comparison...")
        set_progress(0)
        
        try:
            compare_folders_recursive(
                folderA,
                folderB,
                output_dir=output_dir,
                progress_callback=set_progress,
                status_callback=set_status
            )
            messagebox.showinfo("Done", "Comparison complete. Output files created.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

    tk.Button(root, text="Run Comparison", font=("Segoe UI", 11, "bold"), width=20, command=run_comparison).pack(pady=20)

    # -----------------------------
    # Close splash screen AFTER GUI is ready
    # -----------------------------
    if getattr(sys, 'frozen', False) and pyi_splash:
        try:
            pyi_splash.close()
        except Exception:
            pass

    root.mainloop()

def diag(msg):
    import config
          
    #if config.DIAGNOSTIC_MODE:
    print(f"[DIAG] {msg}")


# ============================================================
# 9. ENTRY POINT
# ============================================================

if __name__ == "__main__":

    import argparse
    import config

    parser = argparse.ArgumentParser(
    description="Reconciliation Tool — GUI or CLI mode", formatter_class=argparse.RawTextHelpFormatter)

    parser.add_argument("folderA", nargs="?", help="Source folder (Folder A)")
    parser.add_argument("folderB", nargs="?", help="Target folder (Folder B)")
    parser.add_argument("output_dir", nargs="?", help="Output directory")

    # Comparison flags
    parser.add_argument("--hashonly", action="store_true",help="Use hash-only mode (strict, slowest)")
    parser.add_argument("--hash", action="store_true", help="Use hash comparison (accurate mode)")
    parser.add_argument("--findall", action="store_true", help="Find all matching locations in Folder B")

    # Deletion flags
    parser.add_argument("--deletematches", action="store_true", help="Delete exact matches")
    parser.add_argument("--deletecandidates", action="store_true", help="Delete mismatch candidates")
    parser.add_argument("--noquarantine", action="store_true", help="Disable quarantine folder")
    parser.add_argument("--nodryrun", action="store_true", help="Disable dry-run mode")

    # Output behavior
    parser.add_argument("--timestampedoutput", action="store_true", help="Create a timestamped subfolder inside output_dir")

    # Silent mode
    parser.add_argument("--silent", action="store_true", help="Suppress console output")
    
    # Diagnostic mode
    parser.add_argument("--diagnostic",action="store_true",help="Enable diagnostic output")


    # Comparison Modes help block
    parser.epilog = (
        "Comparison Modes:\n"
        "  (default)          Timestamp + Size only (fastest)\n"
        "  --hash             Timestamp + Size, then hash confirm (accurate)\n"
        "  --hash-only        Hash every file (strict, slowest)\n"
    )
    
    # handle --Output argument
    parser.add_argument(
    "-o", "--output",
    required=False,
    help="Directory where reports will be written",
    )
    # MUST parse args BEFORE using args.output
    # Remember Define Args, Parse Args, Use ArgsF
    args = parser.parse_args()

    output_dir = args.output
    # Defaults output if not passed in.
    if not output_dir: 
        output_dir = os.path.join(os.getcwd(), "reports")

    # --------------------------------------------------------
    # GUI MODE
    # --------------------------------------------------------
    if not args.folderA or not args.folderB:
        launch_gui()
        sys.exit(0)

    # --------------------------------------------------------
    # CLI MODE
    # --------------------------------------------------------
    folderA = args.folderA
    folderB = args.folderB

    # Output directory
    if args.output_dir:
        output_dir = normalize_path(args.output_dir)
    else:
        output_dir = os.getcwd()

    os.makedirs(output_dir, exist_ok=True)

    # --------------------------------------------------------
    # APPLY FLAGS TO GLOBALS
    # --------------------------------------------------------
  
    config.initialize_runtime()

    # Apply CLI arguments to config flags

    # Diagnostics
    if args.diagnostic:
        config.DIAGNOSTIC_MODE = True
        config.DIAG_SCAN = True
        config.DIAG_COMPARE = True
        config.DIAG_WRITERS = True

    # Comparison behavior
    if args.findall:
        config.FIND_ALL_LOCATIONS_MODE = True

    # Hash modes (mutually exclusive)
    if args.hashonly:
        config.HASH_ONLY_MODE = True
        config.HASH_COMPARE_MODE = False

    elif args.hash:
        config.HASH_ONLY_MODE = False
        config.HASH_COMPARE_MODE = True

    # Output behavior
    if args.silent:
        config.SILENT_MODE = True

    # Default is TRUE
    # Timestamped output folder
    if args.timestampedoutput:
        config.TIMESTAMP = True
    
    # Determine output directory
    if args.output:
        output_dir = normalize_path(args.output)
    else:
        output_dir = os.path.join(os.getcwd(), "reports")

    os.makedirs(output_dir, exist_ok=True)
    
    # Deletion / quarantine
    if args.nodryrun:
        config.DRY_RUN = False
    else:
        config.DRY_RUN = True

    if args.noquarantine:
        config.USE_QUARANTINE = False
    else:
        config.USE_QUARANTINE = True

    if args.deletematches:
        config.DELETE_EXACT_MATCHES = True

    if args.deletecandidates:
        config.DELETE_CANDIDATES = True

    # --------------------------------------------------------
    # SILENT MODE CALLBACKS
    # --------------------------------------------------------
    def cli_status(msg):
        if not args.silent:
            print(msg)

    def cli_progress(value):
        pass  # CLI does not need progress bar updates

    # --------------------------------------------------------
    # RUN COMPARISON
    # --------------------------------------------------------
    cli_status("Starting comparison...")

    compare_folders_recursive(
        folderA,
        folderB,
        output_dir=output_dir,
        progress_callback=cli_progress,
        status_callback=cli_status
    )

    cli_status(f"Comparison complete. Output files created in:\n{output_dir}")