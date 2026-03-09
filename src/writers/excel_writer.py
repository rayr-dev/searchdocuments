# writers/excel_writer.py

# System
import os
import logging
from datetime import datetime

# 3rd Party
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Local
from utilities.logging_setup import diag

def write_excel_output(output_dir,
                       matches,
                       mismatched,
                       missing,
                       source_count=0,
                       target_count=0,
                       source_unique=0,
                       target_unique=0,
                       progress_callback=None,
                       status_callback=None
                       ):

    diag("EXCEL_WRITER/write_excel_output Starting")
    excel_path = os.path.join(output_dir, "comparison.xlsx")
    diag(f"Excel Path: {excel_path}")

    # Optional: writer-level status update
    if status_callback:
        status_callback(f"Writing Excel report to: {excel_path}")
    if progress_callback:
        progress_callback(60)
    diag("ENTERING EXCEL WRITER")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"

        # ------------------------------------------------------------
        # Summary counts at top
        # ------------------------------------------------------------
        ws.append(["Total Files in Source:", source_count])
        ws.append(["Total Files in Target:", target_count])
        ws.append(["Unique Filenames Source:", source_unique])
        ws.append(["Unique Filenames Target:", target_unique])
        ws.append([])  # blank row

        header = [
            "Status",
            "Filename",
            "Path A",
            "Timestamp A",
            "Path B",
            "Timestamp B",
            "Action",
        ]
        ws.append(header)

        diag("Checkpoint Exact Matches")

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # ------------------------------------------------------------
        # Exact Matches
        # ------------------------------------------------------------
        for name, pathA, pathB, action in matches:

            if not pathA or not os.path.exists(pathA):
                diag(f"Excel: Skipping invalid pathA for match: {pathA}")
                continue
            if not pathB or not os.path.exists(pathB):
                diag(f"Excel: Skipping invalid pathB for match: {pathB}")
                continue

            try:
                tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsA = ""
            try:
                tsB = datetime.fromtimestamp(os.path.getmtime(pathB)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsB = ""

            row = ["Exact Match", name, pathA, tsA, pathB, tsB, action or ""]
            ws.append(row)

            for cell in ws[ws.max_row]:
                cell.fill = green

        # ------------------------------------------------------------
        # Mismatches
        # ------------------------------------------------------------
        diag("Checkpoint Before Mismatches")

        for name, pathA, b_list, action in mismatched:

            if not pathA or not os.path.exists(pathA):
                diag(f"Excel: Skipping invalid pathA for mismatch: {pathA}")
                continue

            try:
                tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsA = ""

            for pathB, _, _ in b_list:

                if not pathB or not os.path.exists(pathB):
                    diag(f"Excel: Skipping invalid pathB for mismatch: {pathB}")
                    continue

                try:
                    tsB = datetime.fromtimestamp(os.path.getmtime(pathB)).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tsB = ""

                row = ["Mismatch", name, pathA, tsA, pathB, tsB, action or ""]
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.fill = yellow

        # ------------------------------------------------------------
        # Missing Files
        # ------------------------------------------------------------
        diag("Checkpoint Before Missing Files")

        for name, pathA in missing:

            if not pathA or not os.path.exists(pathA):
                diag(f"Excel: Skipping invalid pathA for missing: {pathA}")
                continue

            try:
                tsA = datetime.fromtimestamp(os.path.getmtime(pathA)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                tsA = ""

            row = ["Missing in Folder B", name, pathA, tsA, "", ""]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = red
        diag("Checkpoint Before Save")

        wb.save(excel_path)
        diag("Checkpoint After Save")
    except Exception as error:
        diag(f"ERROR writing Excel report: {error}")
        logging.error(f"ERROR writing Excel report: {error}")
