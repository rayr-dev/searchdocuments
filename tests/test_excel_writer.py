# tests/test_excel_writer.py

import os
import pytest
from unittest.mock import MagicMock
from openpyxl import load_workbook
from src.writers.excel_writer import write_excel_output

# -----------------------------
# Helper fixtures
# -----------------------------
@pytest.fixture
def sample_files(tmp_path):
    """Create real sample files for testing."""
    fileA = tmp_path / "fileA.txt"
    fileB = tmp_path / "fileB.txt"
    fileA.write_text("content A")
    fileB.write_text("content B")
    return str(fileA), str(fileB), str(tmp_path)

def read_xlsx(path):
    """Helper to read xlsx rows as list of lists."""
    wb = load_workbook(path)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]

# -----------------------------
# Basic output tests
# -----------------------------
def test_excel_file_is_created(tmp_path):
    write_excel_output(str(tmp_path), [], [], [])
    assert os.path.exists(str(tmp_path / "comparison.xlsx"))

def test_excel_has_header_row(tmp_path):
    write_excel_output(str(tmp_path), [], [], [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert rows[0] == ["Status", "Filename", "Path A", "Timestamp A", "Path B", "Timestamp B", "Action"]

def test_excel_empty_data_only_header(tmp_path):
    write_excel_output(str(tmp_path), [], [], [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 1  # only header

def test_excel_sheet_title(tmp_path):
    write_excel_output(str(tmp_path), [], [], [])
    wb = load_workbook(str(tmp_path / "comparison.xlsx"))
    assert wb.active.title == "Comparison"

# -----------------------------
# Matches tests
# -----------------------------
def test_excel_writes_exact_match(sample_files):
    fileA, fileB, output_dir = sample_files
    matches = [("file.txt", fileA, fileB, None)]
    write_excel_output(output_dir, matches, [], [])
    rows = read_xlsx(os.path.join(output_dir, "comparison.xlsx"))
    assert len(rows) == 2
    assert rows[1][0] == "Exact Match"
    assert rows[1][1] == "file.txt"

def test_excel_writes_multiple_matches(sample_files):
    fileA, fileB, output_dir = sample_files
    matches = [
        ("file1.txt", fileA, fileB, None),
        ("file2.txt", fileA, fileB, None),
    ]
    write_excel_output(output_dir, matches, [], [])
    rows = read_xlsx(os.path.join(output_dir, "comparison.xlsx"))
    assert len(rows) == 3

def test_excel_skips_match_with_invalid_pathA(tmp_path):
    fileB = tmp_path / "fileB.txt"
    fileB.write_text("content")
    matches = [("file.txt", "fake/path.txt", str(fileB), None)]
    write_excel_output(str(tmp_path), matches, [], [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 1  # only header

def test_excel_skips_match_with_invalid_pathB(tmp_path):
    fileA = tmp_path / "fileA.txt"
    fileA.write_text("content")
    matches = [("file.txt", str(fileA), "fake/pathB.txt", None)]
    write_excel_output(str(tmp_path), matches, [], [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 1  # only header

def test_excel_match_row_has_green_fill(sample_files):
    fileA, fileB, output_dir = sample_files
    matches = [("file.txt", fileA, fileB, None)]
    write_excel_output(output_dir, matches, [], [])
    wb = load_workbook(os.path.join(output_dir, "comparison.xlsx"))
    ws = wb.active
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.rgb == "00C6EFCE"  # green

# -----------------------------
# Mismatches tests
# -----------------------------
def test_excel_writes_mismatch(sample_files):
    fileA, fileB, output_dir = sample_files
    mismatched = [("file.txt", fileA, [(fileB, 100, 1234567890.0)], None)]
    write_excel_output(output_dir, [], mismatched, [])
    rows = read_xlsx(os.path.join(output_dir, "comparison.xlsx"))
    assert len(rows) == 2
    assert rows[1][0] == "Mismatch"

def test_excel_writes_multiple_mismatch_candidates(sample_files):
    fileA, fileB, output_dir = sample_files
    mismatched = [("file.txt", fileA, [
        (fileB, 100, 1234567890.0),
        (fileB, 200, 1234567891.0),
        ],None)]
    write_excel_output(output_dir, [], mismatched, [])
    rows = read_xlsx(os.path.join(output_dir, "comparison.xlsx"))
    assert len(rows) == 3

def test_excel_skips_mismatch_with_invalid_pathA(tmp_path):
    fileB = tmp_path / "fileB.txt"
    fileB.write_text("content")
    mismatched = [("file.txt", "fake/path.txt", [(str(fileB), 100, 0)], None)]
    write_excel_output(str(tmp_path), [], mismatched, [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 1

def test_excel_skips_mismatch_with_invalid_pathB(tmp_path):
    fileA = tmp_path / "fileA.txt"
    fileA.write_text("content")
    mismatched = [("file.txt", str(fileA), [("fake/pathB.txt", 100, 0)], None)]
    write_excel_output(str(tmp_path), [], mismatched, [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 1

def test_excel_mismatch_row_has_yellow_fill(sample_files):
    fileA, fileB, output_dir = sample_files
    mismatched = [("file.txt", fileA, [(fileB, 100, 1234567890.0)], None)]
    write_excel_output(output_dir, [], mismatched, [])
    wb = load_workbook(os.path.join(output_dir, "comparison.xlsx"))
    ws = wb.active
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.rgb == "00FFEB9C"  # yellow

# -----------------------------
# Missing tests
# -----------------------------
def test_excel_writes_missing(sample_files):
    fileA, fileB, output_dir = sample_files
    missing = [("file.txt", fileA)]
    write_excel_output(output_dir, [], [], missing)
    rows = read_xlsx(os.path.join(output_dir, "comparison.xlsx"))
    assert len(rows) == 2
    assert rows[1][0] == "Missing in Folder B"

def test_excel_missing_has_empty_pathB_columns(sample_files):
    fileA, fileB, output_dir = sample_files
    missing = [("file.txt", fileA)]
    write_excel_output(output_dir, [], [], missing)
    rows = read_xlsx(os.path.join(output_dir, "comparison.xlsx"))
    assert rows[1][4] is None or rows[1][4] == ""
    assert rows[1][5] is None or rows[1][5] == ""

def test_excel_skips_missing_with_invalid_pathA(tmp_path):
    missing = [("file.txt", "fake/path.txt")]
    write_excel_output(str(tmp_path), [], [], missing)
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 1

def test_excel_missing_row_has_red_fill(sample_files):
    fileA, fileB, output_dir = sample_files
    missing = [("file.txt", fileA)]
    write_excel_output(output_dir, [], [], missing)
    wb = load_workbook(os.path.join(output_dir, "comparison.xlsx"))
    ws = wb.active
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.rgb == "00FFC7CE"  # red

# -----------------------------
# Timestamp exception tests
# -----------------------------
def test_excel_match_handles_timestamp_error(tmp_path, monkeypatch):
    fileA = tmp_path / "fileA.txt"
    fileB = tmp_path / "fileB.txt"
    fileA.write_text("content")
    fileB.write_text("content")

    def fake_getmtime(path):
        raise OSError("Permission denied")
    monkeypatch.setattr(os.path, "getmtime", fake_getmtime)

    matches = [("file.txt", str(fileA), str(fileB), None)]
    write_excel_output(str(tmp_path), matches, [], [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 2
    assert rows[1][3] == "" or rows[1][3] is None

def test_excel_mismatch_handles_timestamp_error(tmp_path, monkeypatch):
    fileA = tmp_path / "fileA.txt"
    fileB = tmp_path / "fileB.txt"
    fileA.write_text("content")
    fileB.write_text("content")

    def fake_getmtime(path):
        raise OSError("Permission denied")
    monkeypatch.setattr(os.path, "getmtime", fake_getmtime)

    mismatched = [("file.txt", str(fileA), [(str(fileB), 100, 0)], None)]
    write_excel_output(str(tmp_path), [], mismatched, [])
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 2

def test_excel_missing_handles_timestamp_error(tmp_path, monkeypatch):
    fileA = tmp_path / "fileA.txt"
    fileA.write_text("content")

    def fake_getmtime(path):
        raise OSError("Permission denied")
    monkeypatch.setattr(os.path, "getmtime", fake_getmtime)

    missing = [("file.txt", str(fileA))]
    write_excel_output(str(tmp_path), [], [], missing)
    rows = read_xlsx(str(tmp_path / "comparison.xlsx"))
    assert len(rows) == 2

# -----------------------------
# Callback tests
# -----------------------------
def test_excel_status_callback_called(tmp_path):
    status_mock = MagicMock()
    write_excel_output(str(tmp_path), [], [], [], status_callback=status_mock)
    assert status_mock.called

def test_excel_progress_callback_called(tmp_path):
    progress_mock = MagicMock()
    write_excel_output(str(tmp_path), [], [], [], progress_callback=progress_mock)
    progress_mock.assert_called_with(60)

# -----------------------------
# Error handling tests
# -----------------------------
def test_excel_handles_write_error(tmp_path, monkeypatch):
    """Should handle workbook save errors gracefully."""
    from unittest.mock import patch
    with patch("src.writers.excel_writer.Workbook") as mock_wb:
        mock_wb.return_value.save.side_effect = PermissionError("Cannot save")
        write_excel_output(str(tmp_path), [], [], [])

# -----------------------------
# Combined data test
# -----------------------------
def test_excel_writes_all_types(sample_files):
    fileA, fileB, output_dir = sample_files
    matches    = [("match.txt", fileA, fileB, None)]
    mismatched = [("mismatch.txt", fileA,[(fileB,100,1234567890.0)],None)]
    missing    = [("missing.txt", fileA)]
    write_excel_output(output_dir, matches, mismatched, missing)
    rows = read_xlsx(os.path.join(output_dir, "comparison.xlsx"))
    statuses = [r[0] for r in rows[1:]]
    assert "Exact Match" in statuses
    assert "Mismatch" in statuses
    assert "Missing in Folder B" in statuses
